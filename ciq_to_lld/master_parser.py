import openpyxl
from openpyxl import load_workbook
import ipaddress
import sys
import re
from openpyxl.utils import get_column_letter


###### REMOVES QUOTATIONS, COMAS, CURVE BRACES FROM INPUT AND RETURNS VALUE ######
def clean(string):
  temp=str(string).replace("('","").replace("',)","").replace(':','').replace(' ','')
  return temp

def clean_merged_cell(string):
	temp_merged_cell = str(string).replace("<CellRange ","").replace(">","")
	return temp_merged_cell


##### PARSING DATA FROM CIQ AND STORING IN DICT ######
def parse_ciq(aug,nodename):
	CIQ = aug
	CIQwb = load_workbook(CIQ, data_only=True)
	Data = {}
	dict_obj ={}
	modified_key=[]

	for sheets in CIQwb.worksheets:
		if sheets.title in ("IP VLAN"):
			# print("++++"+sheets.title+"++++")
			ws = sheets.title
			CIQws = CIQwb[ws]			
			Data[ws]={}
			key_cell = "C10"
			vallue_cell = "D"
			max_Key_cell = str("C"+str(CIQws.max_row))
			#print(max_Key_cell)
			key_cells = list(openpyxl.utils.rows_from_range(str(key_cell+":"+max_Key_cell)))
			# print(key_cells)
			column_count= CIQws.max_column-3
			for i in range(0,column_count):
				j = 0
				val_start_cell = str(vallue_cell+str(10))
				max_Vallue_cell = str(vallue_cell+str(CIQws.max_row))
				#print(max_Vallue_cell)
				vallue_cells = list(openpyxl.utils.rows_from_range(str(val_start_cell+":"+max_Vallue_cell)))
				header_cell = str(vallue_cell+str(j+9))
				r = CIQws.max_row-9
				while j < r:
					tmp = CIQws[clean(key_cells[j])].value
					temp = CIQws[clean(vallue_cells[j])].value
					if tmp == None:
						j+=3
						header_cell = str(vallue_cell+str(j+9))
					elif temp == None:
						j+=1
					else:						
						z = Data.get(ws,{}).get(CIQws[clean(key_cells[j])].value, 0)
						if z == 0:
							#print(key_cells)
							Data[ws][CIQws[clean(key_cells[j])].value]= {CIQws[clean(header_cell)].value : CIQws[clean(vallue_cells[j])].value}
						else:
							#print(key_cells)
							Data[ws][CIQws[clean(key_cells[j])].value].update({CIQws[clean(header_cell)].value : CIQws[clean(vallue_cells[j])].value})
						j+=1
				
				vallue_cell = chr(ord(vallue_cell)+1)

	for key in Data["IP VLAN"]:		
		modified_key.append(Data["IP VLAN"][key]["Node"].lower() + "_" + key)		
	Data["IP VLAN"] = dict(zip(modified_key, list(Data["IP VLAN"].values())))
	#print(Data["IP VLAN"])
	for key in Data["IP VLAN"]:
		if key.split("_",1)[0] == nodename:
			# print(key.split("_",1)[1])
			
			dict_obj[key.split("_",1)[1]] =Data["IP VLAN"][key]

	# print(dict_obj)
	# for key in dict_obj:
	# 	if key == 'mtas_om1_sp1':	
	# 			print(dict_obj[key])	
	return dict_obj


# parse_ciq("C:\\Users\\elaklak\\Documents\\design_automation\\From Neeraj\\22June2020\\ims-lld-automation\\ciq_to_lld\\Customer_afg.xlsx","afg")
