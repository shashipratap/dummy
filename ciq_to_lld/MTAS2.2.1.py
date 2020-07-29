# -*- coding: utf-8 -*-

"""
MTAS2.2.1.py
~~~~~~~~~~~~

This module implements the MTAS LLD creation for FVL-R2.2 migration. Based on PEP 282 and comments thereto in
comp.lang.python.

:copyright: (c) 2020 by XXXXXXX.All Rights Reserved.
:license: Ericsson , see LICENSE for more details.

class:
Mtas -- responsible for calling  list of
                    resources , and managing the LLD generation mdoules.


    methods:
    
    yaml_file_load() -- It loads the mapping yaml file from predefined path and return a dictionary.
    excel_file_load() -- It helps to load any excel file using openpyxl and returns the workbook object.
    validate_ciq_sheet_by_map() -- It validates whether CIQ defined sheetnames are present in yaml mapping and returns list of matched sheetnames.
    create_mtas_lld() -- It extracts and excutes both excel file and yaml files together and creates the Final LLD for specific nodes.
    create_ip_sheet() -- It extracts and excutes both excel file and yaml files together and creates the Final LLD for IP VLAN Sheet.
    redundant_param_name() -- 
    cell_color() --
    ip_sheet_lookup() --
    border() --
    special_cases() --
    clean_redundant_column() --

"""

# Import Dependancies
import yaml
import openpyxl
import ipaddress
from math import pow
from netaddr import *
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Color, Fill, Font, PatternFill
#from Core.CustomLogger import  getLogger
from openpyxl.utils.cell import coordinate_from_string,column_index_from_string
from master_parser import parse_ciq 
from openpyxl.styles.borders import Border, Side, BORDER_THIN



#log = getLogger('root')

__author__  = "Chandrayee Kumar <chandrayee.kumar@ericsson.com> Neeraj J Sharma <neeraj.j.sharma@ericsson.com>"
__status__  = "development"
# The following module attributes are no longer updated.
__version__ = "2.2.1"
__date__    = "06 June 2020"

class Mtas(object):
    """Raised when LLD creation for MTAS requests invokes."""
    def __init__(self):
        # self.ciq_path="ciq_to_lld/Mobily Saudi_vMTAS_updated_site1.xlsx"
        self.ciq_path="ciq_to_lld/Customer_mtas.xlsx"
        # self.master_path="C:/Users/ekcuhma/OneDrive - Ericsson AB/Local/design automation/ims-lld-automation/ciq_to_lld/Master Sheet.xlsx"
        # self.output_path="ciq_to_lld/Mobily Saudi_vMTAS_Final_LLD_site1.xlsx"
        self.output_path="output/Final_Mtas_lld.xlsx"
        self.mapping_path="Data/MTAS.yml"
        self.template_path="ciq_to_lld/FVL R2.2 - vMTAS 1.17.0 LLD_Mapping version-PA2(Neeraj).xlsx"
        self.matched_sheet_name =[] 
        self.second_column=[]
        self.source_redundant_column = []
        self.ip_list=[]
        self.flag=False
        self.value =0
        self.VIP_cell_list = []
        self.v6_list = []
        self.resource_dict = self.yaml_file_load(self.mapping_path)
        self.fvl_ciq_wb = self.excel_file_load(self.ciq_path)
        self.fvl_lld_wb = self.excel_file_load(self.template_path)
        if self.resource_dict.get("special", None):
            self.second_column = self.redundant_param_name(self)
            self.source_redundant_column = self.redundant_param_name_source(self)
        # for index in self.second_column:
        #     print(len(index))
        self.ip_lookup = {}
       # log.info('message from SubClass / __init__')

    def __del__(self):
        classname = self.__class__.__name__
       # log.info(classname +" deleted")   
    
    def cell_color(self, worksheet, cell,bgcolor,fontcolor):
        """Adds color to updated cells"""
        # bgcolor="openpyxl.styles.colors."+bgcolor
        # fontcolor="openpyxl.styles.colors."
        color_fill = PatternFill(end_color=bgcolor,fill_type='solid',start_color=bgcolor)
        worksheet[cell].fill = color_fill
        worksheet[cell].font = Font(color=fontcolor)

    def border(self):
        border = Border(left=Side(border_style=BORDER_THIN,color='FF000000'),right=Side(border_style=BORDER_THIN,color='FF000000'),top=Side(border_style=BORDER_THIN, color='FF000000'),bottom=Side(border_style=BORDER_THIN,color='FF000000'))   
        return border
    
    def redundant_param_name(self,*args):
        yaml_obj = self.resource_dict['Special']['Data_Redundancy']['Destination']
        p_list =[]
        for i in range(0,len(yaml_obj)):
            p_list.append([x for x in self.fvl_lld_wb[yaml_obj[i]['sheetname']][yaml_obj[i]['column']] if yaml_obj[i]['param'] == x.value])
            # print([x for x in self.fvl_lld_wb[yaml_obj[i]['sheetname']][yaml_obj[i]['column']] if yaml_obj[i]['param'] == x.value])
        return p_list
    
    def redundant_param_name_source(self,*args):
        yaml_obj = self.resource_dict['Special']['Data_Redundancy']['Source']
        source_list =[]
        for i in range(0,len(yaml_obj)):
            source_list.append([x for x in self.fvl_ciq_wb[yaml_obj[i]['sheetname']][yaml_obj[i]['column']] if yaml_obj[i]['param'] == x.value])
            # print([x for x in self.fvl_ciq_wb[yaml_obj[i]['sheetname']][yaml_obj[i]['column']] if yaml_obj[i]['param'] == x.value])
        return source_list    
    
    def ip_sheet_lookup(self,*args,**kwargs):
        # print(args[0])
        switcher = {
            # Final LLD:Customer CiQ sheet
            # 'Information':'Information',
            'Vlan Name':'Vlan Name',
            'VPN':'VPN',
            'Protocol/Mask':'Subnet size',
            'VLANID':'VLAN ID assigned by',
            'Node':'Node',
            'IPv4 Address':'Assigned IP(IPv4) range by',
            # 'Comment':'Comment',
            'IPv6 Address':'Assigned IP(Ipv6) range by',
            'Logical Connectivity Description':'Logical Connectivity Description'
                    }    
        resp = switcher.get(args[0], lambda: "Invalid Column Name")
        # print("REEEEE",resp)
        return resp
        
    def yaml_file_load(self,*args,**kwargs):
          
        """
               Sends a fetch request, but retrieves the
               resource_dict.Returns :dictionary :`resource_dict` object.
        
               :param args[0]: yaml file location (i.e.: self.mapping_path)
               :param **kwargs: Optional arguments that ``request`` takes.
        """

        resource_dict = yaml.safe_load(open(args[0],'r',encoding='utf-8-sig'))
        return resource_dict

    def excel_file_load(self,*args,**kwargs):
        """
               Sends a fetch request, but retrieves the
               workbook object.Returns :Object :`wb` object.
        
               :param args[0]: excel file location (i.e.: self.ciq_path,self.output_path)
               :param **kwargs: Optional arguments that ``request`` takes.
        """
        wb = load_workbook(args[0])
        # print('Inside',wb.sheetnames)
        return wb

    def validate_ciq_sheet_by_map(self,*args,**kwargs): 
        """
               Sends a search request, but retrieves the
               list of sheet object.Returns :List :`matched_sheet_name` object.
        
               :param args[0]: workbook object of CIQ (i.e.:  self.fvl_ciq_wb)
                               Dictionary object of yaml file (i.e.:  self.resource_dict)
               :param **kwargs: Optional arguments that ``request`` takes.
        """
        
        # sheet_list=self.fvl_ciq_wb.sheetnames
        sheet_list=[x.strip(' ') for x in self.fvl_ciq_wb.sheetnames]
        # print("fvl_ciq_wb",self.resource_dict['Generic'])
        
        for ciq_sheet_name in self.resource_dict['Generic']:
            # print("self.resource_dict['Generic']",ciq_sheet_name.strip())
            if ciq_sheet_name.strip() in sheet_list:
                # print("Matched",type(ciq_sheet_name))
                self.matched_sheet_name.append(ciq_sheet_name)
            else:
                #log.error("Please check your mapping file SHEET name with CIQ file SHEET name")
                #log.error("Error Code:")
                # print("Non Matched",ciq_sheet_name) 
                pass      
        return self.matched_sheet_name

    def create_mtas_lld(self,*args,**kwargs):
        """
               Create a get request, but retrieves the
               list of sheet object.Returns :List :`matched_sheet_name` object.
        
               :param args[0]: Common List of workbook of CIQ and yaml (i.e.:  args[0].matched_sheet_name)
                               path object of CIQ file (i.e.:  args[0].ciq_path)
                               path object of yaml file (i.e.:  args[0].mapping_path)
                               path object of LLD file (i.e.:  args[0].output_path)
               :param **kwargs: Optional arguments that ``request`` takes.
        """
        matched_sheet_name = args[0].matched_sheet_name
        ciq_path = args[0].ciq_path
        mapping_path = args[0].mapping_path
        matched_sheet_name=self.validate_ciq_sheet_by_map(mapping_path,ciq_path)
        # print(self.resource_dict['Generic'])
        ### here using the sheet generated by customer_name module
        for ciq_sheet_name in tqdm(matched_sheet_name):
            # print(ciq_sheet_name)
            sheet_obj=self.fvl_ciq_wb[ciq_sheet_name]
            for ciq_param_name in self.resource_dict['Generic'][ciq_sheet_name]:
                # print("SHEET NAME",ciq_sheet_name)
            #   if ciq_sheet_name == 'ICS':
                for row in sheet_obj.iter_rows(min_col=2, min_row=sheet_obj.min_row, max_col=2, max_row=sheet_obj.max_row):
                        for cell in row:
                            for index in self.source_redundant_column:
                                for (val,i)  in zip(index,range(1,len(index)+1)):            
                                    if '@' not in val.value and val.value == cell.value:
                                        cell.value = str(val.value)+"@"+str(i)
                                        # print(cell.value)
                            # log.info(ciq_param_name ,"==", cell.value)
                            if ciq_param_name == cell.value:
                                # log.info("when YAML == ciq value",cell.value)
                                for lld_sheet_name in  self.resource_dict['Generic'][ciq_sheet_name][ciq_param_name]:
                                    for lld_param_name in self.resource_dict['Generic'][ciq_sheet_name][ciq_param_name][lld_sheet_name]:
                                        lld_sheet_obj=self.fvl_lld_wb[lld_sheet_name] 
                                        # if self.resource_dict['Special']['Data_Redundancy'].Contains(lld_sheet_name):
                                        # print(self.second_column)
                                        for index in self.second_column:
                                            # print(index)
                                            for (val,i)  in zip(index,range(1,len(index)+1)):
                                                # print(lld_sheet_obj[val.coordinate],"########",val)   
                                                if '%' not in val.value and self.flag == False and lld_sheet_name in str(val):
                                                    lld_sheet_obj[val.coordinate].value = str(val.value)+"%"+str(i) 
                                                    # self.cell_color(lld_sheet_obj,val.coordinate)
                                                
                                            # if lld_sheet_name == 'MTAS - SMTASSharedIFC':
                                           
                                                 
                                        for row1 in lld_sheet_obj.iter_rows(min_col=2, min_row=lld_sheet_obj.min_row,max_col=2, max_row=lld_sheet_obj.max_row):
                                                # print(row1,row1.count(lld_param_name))
                                                for cell1 in row1:
                                                    # print(type(row1))
                                                    # print(coordinate_from_string(cell1.coordinate))
                                                    # print(lld_param_name, '==' ,cell1.value)
                                                    # redList =[]
                                                    if lld_param_name == cell1.value:
                                                        # print(lld_param_name, '==' ,cell1.value)    
                                                            
                                                        for i in self.resource_dict['Generic'][ciq_sheet_name][ciq_param_name][lld_sheet_name][lld_param_name]:
                                                            ciq_param_col,lld_param_col = i.split(":")
                                                            
                                                            if '%' in str(lld_sheet_obj[cell1.coordinate].value):
                                                                # print("BEFORE",str(lld_sheet_obj[cell1.coordinate].value))
                                                                lld_sheet_obj[cell1.coordinate].value = lld_sheet_obj[cell1.coordinate].value.split("%")[0]
                                                                # self.cell_color(lld_sheet_obj,cell.coordinate,'9400D3','ffffff')
                                                                # self.cell_color(lld_sheet_obj,cell1.coordinate)
                                                                # print("AFTER",lld_sheet_obj[cell1.coordinate].value)
                                                                self.flag=True
                                                            # print(str(lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value))
                                                            if '<ip>' in str(lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value):     
                                                                lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value = str(sheet_obj[cell.coordinate.replace(cell.coordinate[0],ciq_param_col)].value).split(":")[0]
                                                            #purple color codinf for testing
                                                                print(cell1.coordinate.replace(cell1.coordinate[0],lld_param_col))
                                                                self.cell_color(lld_sheet_obj,lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].coordinate,'9400D3','ffffff')
                                                                print(lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value,lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)])
                                                                self.ip_list.append(cell1.coordinate)
                                                            if '<port>' in str(lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value):     
                                                                lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value = str(sheet_obj[cell.coordinate.replace(cell.coordinate[0],ciq_param_col)].value).split(":")[1]
                                                            #purple color codinf for testing
                                                                print(cell1.coordinate.replace(cell1.coordinate[0],lld_param_col))
                                                                self.cell_color(lld_sheet_obj,lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].coordinate,'9400D3','ffffff')
                                                                print(lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value,lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)])
                                                                self.ip_list.append(cell1.coordinate) 
                                                            if '<input>' in str(lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value):     
                                                                lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value = lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value.replace("<input>",str(sheet_obj[cell.coordinate.replace(cell.coordinate[0],ciq_param_col)].value))
                                                                
                                                            #purple color codinf for testing
                                                                print(cell1.coordinate.replace(cell1.coordinate[0],lld_param_col))
                                                                self.cell_color(lld_sheet_obj,lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].coordinate,'9400D3','ffffff')
                                                                print(lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value,lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)])
                                                                self.ip_list.append(cell1.coordinate)       
                                                            elif cell1.coordinate not in str(self.ip_list):
                                                                lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].value = sheet_obj[cell.coordinate.replace(cell.coordinate[0],ciq_param_col)].value
                                                            #purple color codinf for testing
                                                                self.cell_color(lld_sheet_obj,lld_sheet_obj[cell1.coordinate.replace(cell1.coordinate[0],lld_param_col)].coordinate,'9400D3','ffffff')

                                                            # self.cell_color(lld_sheet_obj,cell1.coordinate)
                                                    else:
                                                        pass
                                                            # print(cell1.value,cell1.coordinate)

                                                        # log.error("Please check your mapping file parameter name with lld file parameter name")
                                                        # log.error("Error Code:")
                                        # print(redList)
                                        # for row1 in lld_sheet_obj.iter_rows(min_col=2, min_row=lld_sheet_obj.min_row,max_col=2, max_row=lld_sheet_obj.max_row):
                                        #         # print(row1,row1.count(lld_param_name))
                                        #         for cell1 in row1:
                                        
                                        #             if '%' in str(cell1.value):
                                        #                                 lld_sheet_obj[cell1.coordinate].value = lld_sheet_obj[cell1.coordinate].value.split("%")[0]              
                                            
                                            # self.fvl_lld_wb.save(self.output_path)
                                        
                            
                            else:
                                                    pass
                                                    # log.error("Please check your mapping file parameter name with CIQ file parameter name")
                                                    # log.error("Error Code:") 
                                                    
                                                    
        
        self.fvl_lld_wb.save(self.output_path)  
      
    def create_ip_sheet(self,*args,**kwargs):
        """
               Create a get request, but retrieves the
               list of sheet object.Returns :List :`matched_sheet_name` object.
        
               :param args[0]: Common List of workbook of CIQ and yaml (i.e.:  args[0].matched_sheet_name)
                               path object of CIQ file (i.e.:  args[0].ciq_path)
                               path object of yaml file (i.e.:  args[0].mapping_path)
                               path object of LLD file (i.e.:  args[0].output_path)
               :param **kwargs: Optional arguments that ``request`` takes.
        """
        master_dict=parse_ciq(self.ciq_path,'mtas')
        self.fvl_lld_wb_ip = self.excel_file_load(self.output_path)  
        ip_sheet_obj=self.fvl_lld_wb_ip['vMTAS IP Plan']
        # print(ip_sheet_obj)
        min_row=20      
        min_row_next=20
        min_row_ipv6=20
        for key in master_dict:
            # print('KEYY',key)
            for inner_key in master_dict[key]:
                # print("INN",inner_key)
                #removed if condition 'VIP' not in key and 
                
                if 'VIP'.casefold() not in key and 'IPv4' in inner_key and "/" in master_dict[key][inner_key] and 'e.g.:' not in master_dict[key][inner_key]:
                    # print("ORG",master_dict[key][inner_key].split("/")[1])
                    
                    
                    vlan_count = int((pow(2,32 - int(master_dict[key][inner_key].split("/")[1]))))-1
                    # print("Iertate",vlan_count)
                    vlan_flag = False
                    if vlan_count > 8:
                        vlan_count_long=vlan_count
                        vlan_count = 14
                    # self.VIP_cell_list[master_dict[key][inner_key ]]= vlan_count
                    for row in ip_sheet_obj.iter_rows(min_col=6, min_row=min_row, max_col=6, max_row=min_row+vlan_count):
                        for cell in row:
                            # print(key)
                            cell.value = key
                            cell.border = self.border()
                            
                    address=master_dict[key][inner_key].split("/")[0]  
                    counter = 0      
                    for row in ip_sheet_obj.iter_rows(min_col=2, min_row=min_row, max_col=2, max_row=min_row+vlan_count):
                        for cell in row:
                            counter +=1
                            if counter == 1:
                                cell.value = str(master_dict[key][inner_key])
                                cell.border = self.border()
                                v=coordinate_from_string(cell.coordinate)[1]
                                # print(master_dict[key])
                                for i in master_dict[key]:
                                    if "Ipv6" in i and "/" in master_dict[key][i]:
                                        # print(v)
                                        self.v6_list.append(v)
                                for row in ip_sheet_obj.iter_rows(min_col=2, min_row=v, max_col=10, max_row=v):
                                    for cell in row:
                                        self.cell_color(ip_sheet_obj,cell.coordinate,'ffcc00','000000')
                                        cell.border = self.border()
                                # self.cell_color(ip_sheet_obj,cell)
                            elif counter == 12:
                                cell.value = str("...") 
                                cell.border = self.border()   
                            elif counter == 13:
                                # print(ipaddress.ip_address(address),vlan_count_long,vlan_count)
                                cell.value = str(ipaddress.ip_address(master_dict[key][inner_key].split("/")[0]) + int(vlan_count_long) - 2)
                                cell.border = self.border()
                            elif counter == 14:
                                # print(ipaddress.ip_address(address),vlan_count_long,vlan_count)
                                cell.value = str(ipaddress.ip_address(master_dict[key][inner_key].split("/")[0]) + int(vlan_count_long) - 1)
                                cell.border = self.border()
                            elif counter == 15:
                                # print(ipaddress.ip_address(address),vlan_count_long,vlan_count)
                                cell.value = str(ipaddress.ip_address(master_dict[key][inner_key].split("/")[0]) + int(vlan_count_long))     
                                cell.border = self.border()
                            # elif counter == 16:
                            #     # print(ipaddress.ip_address(address),vlan_count_long,vlan_count)
                            #     cell.value = str(ipaddress.ip_address(master_dict[key][inner_key].split("/")[0]) + int(vlan_count_long))
                            else:
                                address = ipaddress.ip_address(address) + 1
                                # print(cell.coordinate,address)
                                cell.value = str(address)
                                cell.border = self.border()
                    
                    min_row=min_row+vlan_count+1    
                   
                elif 'VIP'.casefold()  in key:
                    # print(key)
                    self.VIP_cell_list.append(key)

                for min_row_ipv6 in self.v6_list:    
                    if  'Ipv6' in inner_key and  "/" in master_dict[key][inner_key] and 'e.g.:' not in master_dict[key][inner_key]:
                            # print(master_dict[key][inner_key],"@@@@@@@@",IPNetwork(master_dict[key][inner_key]).size)
                            
                            if IPNetwork(master_dict[key][inner_key]).size > 1:
                                    # print("ORG",IPNetwork(master_dict[key][inner_key])[0])
                                    vlan_count = 14
                            else:
                                    vlan_count =  IPNetwork(master_dict[key][inner_key]).size      
                            # for ip in IPNetwork(master_dict[key][inner_key]):
                        
                            # print("COUNT",min_row_ipv6,min_row_ipv6+vlan_count) 
                            if master_dict[key][inner_key] is not None :
                                ip_in_network = IPNetwork(master_dict[key][inner_key])
                                counter_ip = 0
                                
                                j=5
                            
                                for row in ip_sheet_obj.iter_rows(min_col=3, min_row=min_row_ipv6, max_col=3, max_row = min_row_ipv6+vlan_count):
                                    # for i in range(1,vlan_count):
                                            for cell in row:
                                                        
                                                        if counter_ip == 0:
                                                        
                                                            cell.value= str(master_dict[key][inner_key])
                                                            cell.border = self.border()
                                                            # print(cell.coordinate,cell.value)
                                                        elif counter_ip > 0 and counter_ip < 11:
                                                            # print(counter_ip,"INNNN")
                                                            cell.value=str(ip_in_network[counter_ip])
                                                            cell.border = self.border()
                                                            # print(cell.coordinate,cell.value)
                                                        elif counter_ip == 11 :
                                                            cell.value="..."
                                                            cell.border = self.border()
                                                            # print(cell.coordinate,cell.value)
                                                        else:
                                                            j=j-1
                                                            # print(j,"ggggggggg")
                                                            cell.value= str(ip_in_network[-(j)])
                                                            cell.border = self.border()
                                                            # print(cell.coordinate,cell.value)
                                                        counter_ip+=1
                            # min_row_ipv6 =min_row_ipv6+vlan_count+1     
            new_list = list(set(list(self.VIP_cell_list)))
            for row in ip_sheet_obj.iter_rows(min_col=3, min_row=min_row_next, max_col=10, max_row=min_row-1):
                        
                        for inner_key in master_dict[key]:
                            for cell in row:
                                count = 'F'+str(min_row_next)
                                # print(key)
                                if 'F' not in cell.coordinate and str(self.ip_sheet_lookup(ip_sheet_obj[coordinate_from_string(cell.coordinate)[0]+str('4')].value)) in inner_key :
                                    # print(str(self.ip_sheet_lookup(ip_sheet_obj[coordinate_from_string(cell.coordinate)[0]+str('4')].value)),"$$$$$", inner_key)
                                    cell.value = master_dict[key][inner_key]
                                    cell.border = self.border()
                                else:
                                    cell.border = self.border()    

                        min_row_next +=1

        for (row,i) in zip(ip_sheet_obj.iter_rows(min_col=2, min_row=min_row_next+1, max_col=10, max_row=min_row_next+len(list(set(self.VIP_cell_list)))),new_list): 
            for key in master_dict:
                if  i == key :     
                    for inner_key in master_dict[key]:            
                        for cell in row :
                            if str(self.ip_sheet_lookup(ip_sheet_obj[coordinate_from_string(cell.coordinate)[0]+str('4')].value)) in inner_key :
                                            if '/32' and '.' in str(master_dict[key][inner_key]):
                                                    # print(master_dict[key][inner_key].split('/')[0])
                                                    cell.value = master_dict[key][inner_key].split('/')[0]
                                            else:
                                                cell.value = master_dict[key][inner_key]
                                            cell.border = self.border()
                                            # print(cell.coordinate,cell.value)
                            else:
                                            cell.border = self.border()                
        for row in ip_sheet_obj.iter_rows(min_col=2, min_row=min_row_next, max_col=10, max_row=min_row_next):
            for cell in row :
                            if str(min_row_next) in cell.coordinate:
                                if 'C' in cell.coordinate:
                        # print(cell.coordinate)
                                    cell.value = " VIP Address "
                                    cell.border = self.border()
                                else :    
                                    cell.value = " "
                                self.cell_color(ip_sheet_obj,cell.coordinate,'00004d','ffffff')  
                                cell.border = self.border()          
        self.fvl_lld_wb_ip.save(self.output_path)

    def special_cases(self,*args,**kwargs):
        """
               Create a get request, but retrieves the
               list of sheet object.Returns :List :`matched_sheet_name` object.
        
               :param args[0]: Common List of workbook of CIQ and yaml (i.e.:  args[0].matched_sheet_name)
                               path object of CIQ file (i.e.:  args[0].ciq_path)
                               path object of yaml file (i.e.:  args[0].mapping_path)
                               path object of LLD file (i.e.:  args[0].output_path)
               :param **kwargs: Optional arguments that ``request`` takes.
        """        
        #print("arg[0] is ",args[0])
        matched_sheet_name = args[0].matched_sheet_name
        #print(matched_sheet_name)
        #sys.exit()
        output_path = args[0].output_path
        ciq_path = args[0].ciq_path
        mapping_path = args[0].mapping_path
        fvl_output_wb = load_workbook(output_path)
        ciq_sheet_list = self.fvl_ciq_wb.sheetnames
        lld_sheet_list = fvl_output_wb.sheetnames
        
        # print(self.resource_dict['Generic'])
        ### here using the sheet generated by customer_name module
        for key in self.resource_dict["Special"]:
            if(key == "Composite"):
                composite_param_flag = 0
                for sheet in self.resource_dict["Special"][key]:                    
                    if  sheet in ciq_sheet_list:
                        fvl_ciq_ws = self.fvl_ciq_wb[sheet]
                        # print(sheet)
                    for param in self.resource_dict["Special"][key][sheet]:
                        # if(len(param.split(",")>1)):                            
                        #     composite_param_flag = 1
                        # else:                            
                        #     composite_param_flag = 0
                        maxcol = fvl_ciq_ws.max_column
                        maxrow = fvl_ciq_ws.max_row
                        composite_ciq_param_value_list = []
                        #print(param.split(","))                                    
                        for i in range(len(param.split(","))):
                            # print("here",i)
                            for col in fvl_ciq_ws.iter_cols(min_row=0, min_col=0, max_row=maxrow , max_col = maxcol):
                                for cell in col:
                                    index=cell.coordinate
                                    cell_index=openpyxl.utils.cell.coordinate_from_string(index)                            
                                    for j in range(0,1):
                                        ciq_col_value=cell_index[0]
                                        ciq_row_value=cell_index[1]
                                    if ciq_col_value == "B":
                                    #composite_ciq_param_value_list = []                                    
                                    #for i in range(len(param.split(","))):
                                        # print(param.split(",")[i], i) 
                                        if(param.split(",")[i] == cell.value):
                                            for sheet_name_lld in self.resource_dict["Special"][key][sheet][param]:
                                                fvl_output_ws = fvl_output_wb[sheet_name_lld] 									
                                                maxrow_lld = fvl_output_ws.max_row
                                                maxcol_lld = fvl_output_ws.max_column
                                
                                                for parameter_lld in self.resource_dict["Special"][key][sheet][param][sheet_name_lld]:									
                                        
                                                    for column_info in self.resource_dict["Special"][key][sheet][param][sheet_name_lld][parameter_lld]:											
                                                        #if(len(column_info.split(",",1)) >1):
                                                        ciq_column,lld_column = column_info.split(":",1)
                                                        if(len(ciq_column.split(",",1)) >1):
                                                            ciq_mapping_value = fvl_ciq_ws[(ciq_column.split(",",1)[0]) + str(ciq_row_value)].value + "," + fvl_ciq_ws[ciq_column.split(",",1)[1] + str(ciq_row_value)].value
                                                            # print("ciq_mapping_value, ",ciq_mapping_value)
                                                        else:                                                             
                                                            ciq_mapping_value = fvl_ciq_ws[ciq_column + str(ciq_row_value)].value
                                                            composite_ciq_param_value_list.append(ciq_mapping_value)
                                                            # print(composite_ciq_param_value_list)
                                            
                                                    for col_lld in fvl_output_ws.iter_cols(min_row=1, min_col=2, max_row= maxrow_lld, max_col=maxcol_lld):
                                                
                                                        for cell_lld in col_lld:
                                                            index_lld=cell_lld.coordinate
                                                            cell_index_lld=openpyxl.utils.cell.coordinate_from_string(index_lld)
                                                    
                                                            for i in range(0,1):
                                                                lld_col_value=cell_index_lld[0]
                                                                lld_row_value=cell_index_lld[1]                                                    
                                                            if(cell_lld.value == parameter_lld ):														
                                                                # print("True")
                                                                absolute_cell_lld = lld_column + str(lld_row_value)														
                                                                # print(cell_lld.value)
                                                                self.cell_color(fvl_output_ws, absolute_cell_lld,'9400D3','ffffff')
                                                                if(len(composite_ciq_param_value_list)>1):
                                                                    ciq_value = composite_ciq_param_value_list[0] +"," + composite_ciq_param_value_list[1]
                                                                    fvl_output_ws[absolute_cell_lld].value = ciq_value
                                                                else:
                                                                    fvl_output_ws[absolute_cell_lld].value  = ciq_mapping_value
                                        
                
                fvl_output_wb.save(self.output_path)                     
   
    def clean_redundant_column(self,*args,**kwargs):
        """
               Create a get request, but retrieves the
               list of sheet object.Returns :List :`matched_sheet_name` object.
        
               :param args[0]: Common List of workbook of CIQ and yaml (i.e.:  args[0].matched_sheet_name)
                               path object of CIQ file (i.e.:  args[0].ciq_path)
                               path object of yaml file (i.e.:  args[0].mapping_path)
                               path object of LLD file (i.e.:  args[0].output_path)
               :param **kwargs: Optional arguments that ``request`` takes.
        """        
        self.fvl_lld_wb_clean = self.excel_file_load(self.output_path) 
        yaml_obj = self.resource_dict['Special']['Data_Redundancy']['Destination']
        for i in range(0,len(yaml_obj)):
            for index in list(self.fvl_lld_wb_clean[yaml_obj[i].get('sheetname')][yaml_obj[i].get('column')]):
                if '%' in str(index.value):
                    index.value= index.value.split("%")[0]
                    
        self.fvl_lld_wb_clean.save(self.output_path)

    def same_sheet_data(self,*args,**kwargs):
        ciq_data = parse_ciq(self.ciq_path, "mtas")
        print(ciq_data)
        # lld_path = self.special_cases(args[0])   
        lld_path = self.output_path
        print(lld_path)
        lld_wb = load_workbook(lld_path)
        lld_sheet_list = lld_wb.sheetnames
        #print(lld_sheet_list)
        for key in self.resource_dict["Same Sheet Data"]:
            if(key == "General"):
                #print(key)
                for sheet in self.resource_dict["Same Sheet Data"][key]:
                    if(sheet in lld_sheet_list):
                        lld_ws = lld_wb[sheet]
                    for param in self.resource_dict["Same Sheet Data"][key][sheet]:
                        maxcol = lld_ws.max_column
                        maxrow = lld_ws.max_row
                        for col in lld_ws.iter_cols(min_row=0, min_col=0, max_row=maxrow , max_col = maxcol):
                            for cell in col:
                                index=cell.coordinate
                                cell_index=openpyxl.utils.cell.coordinate_from_string(index)                            
                                for j in range(0,1):
                                    lld_col_value=cell_index[0]
                                    lld_row_value=cell_index[1]
                                if lld_col_value == "B":
                                    #print(param_sheet,cell.value)
                                    if(cell.value == param):
                                        for inner_key in self.resource_dict["Same Sheet Data"][key][sheet][param]:
                                            lld_ws1=lld_wb[inner_key]
                                            inner_cell_coord,cell_coord = self.resource_dict["Same Sheet Data"][key][sheet][param][inner_key]
                                            if(param == "sipUri" and sheet == "Emergency"):
                                                lld_ws[cell_coord].value = "911@" + lld_ws1[inner_cell_coord].value
                                                self.cell_color(lld_ws, cell_coord,'9400D3','ffffff')
                                            else:    
                                                lld_ws[cell_coord].value = lld_ws1[inner_cell_coord].value
                                                self.cell_color(lld_ws, cell_coord,'9400D3','ffffff')
                lld_wb.save(lld_path)
                               
            elif(key == "IP"):
                #print(key)
                for sheet in self.resource_dict["Same Sheet Data"][key]:
                    if(sheet in lld_sheet_list):
                        lld_ws = lld_wb[sheet]
                    for param_sheet in self.resource_dict["Same Sheet Data"][key][sheet]:
                        maxcol = lld_ws.max_column
                        maxrow = lld_ws.max_row
                        for col in lld_ws.iter_cols(min_row=0, min_col=0, max_row=maxrow , max_col = maxcol):
                            for cell in col:
                                index=cell.coordinate
                                cell_index=openpyxl.utils.cell.coordinate_from_string(index)                            
                                for j in range(0,1):
                                    lld_col_value=cell_index[0]
                                    lld_row_value=cell_index[1]
                                if lld_col_value == "B":
                                    #print(param_sheet,cell.value)
                                    if(cell.value == param_sheet):
                                        print(self.resource_dict["Same Sheet Data"][key])                                        
                                        vlan_type,ip_location,cell_coord = self.resource_dict["Same Sheet Data"][key][sheet][param_sheet]                                        
                                        print(vlan_type,ip_location,cell_coord)
                                        #sys.exit()
                                        for vlan_key in ciq_data:
                                            if "Vlan Type" in ciq_data[vlan_key]: 
                                                print(ciq_data[vlan_key]["Vlan Type"])                                                         
                                                if((ciq_data[vlan_key]["Vlan Type"]) == vlan_type):
                                                    if(ip_location == "Gateway IPv4"):
                                                        ip_list = IPNetwork(ciq_data[vlan_key]["Assigned IP(IPv4) range"])
                                                        lld_ws[cell_coord].value = self.clean_IPaddress(str(ip_list[1]))
                                                        self.cell_color(lld_ws, cell_coord,'9400D3','ffffff')
                                                    elif(ip_location == "Gateway IPv6"):
                                                        if("Assigned IP(Ipv6) range" not in ciq_data[vlan_key]):
                                                            lld_ws[cell_coord].value = "0::0"
                                                            self.cell_color(lld_ws, cell_coord,'9400D3','ffffff')
                                                        else:
                                                            ipv6_list = IPNetwork(ciq_data[vlan_key]["Assigned IP(IPv6) range"])
                                                            lld_ws[cell_coord].value = self.clean_IPaddress(str(ipv6_list[1]))
                                                            self.cell_color(lld_ws, cell_coord,'9400D3','ffffff')
                                                    elif(ip_location == "Subnet Mask"):
                                                        lld_ws[cell_coord].value = str((ciq_data[vlan_key]["Subnet size"]))                          
                                                        self.cell_color(lld_ws, cell_coord,'9400D3','ffffff')
                                                    elif(ip_location == "VLAN ID"):
                                                        lld_ws[cell_coord].value = str((ciq_data[vlan_key]["VLAN ID"]))                          
                                                        self.cell_color(lld_ws, cell_coord,'9400D3','ffffff')
                                                    elif(ip_location == "Mated Pair IP"):
                                                        ip_list = IPNetwork(ciq_data[vlan_key]["Assigned IP(IPv4) range"])
                                                        lld_ws[cell_coord].value = self.clean_IPaddress(str(ip_list[4]))
                                                        self.cell_color(lld_ws, cell_coord,'9400D3','ffffff')
                lld_wb.save(lld_path)                                            
                #return(lld_path)      
                
            elif(key == "VIP"):
                #print(key)
                for sheet in self.resource_dict["Same Sheet Data"][key]:
                    if(sheet in lld_sheet_list):
                        lld_ws = lld_wb[sheet]
                    for param_sheet in self.resource_dict["Same Sheet Data"][key][sheet]:
                        maxcol = lld_ws.max_column
                        maxrow = lld_ws.max_row
                        for col in lld_ws.iter_cols(min_row=0, min_col=0, max_row=maxrow , max_col = maxcol):
                            for cell in col:
                                index=cell.coordinate
                                cell_index=openpyxl.utils.cell.coordinate_from_string(index)                            
                                for j in range(0,1):
                                    ciq_col_value=cell_index[0]
                                    ciq_row_value=cell_index[1]
                                if ciq_col_value == "B":
                                    #print(param_sheet,cell.value)
                                    if(cell.value == param_sheet):
                                        for inner_key in self.resource_dict["Same Sheet Data"][key][sheet][param_sheet]:
                                            vip_type, cell = inner_key.split(",",1)
                                            print(vip_type,cell)
                                            for vip_key in ciq_data:
                                                if "Vip Type" in ciq_data[vip_key]: 
                                                    print((ciq_data[vip_key]["Vip Type"]).split("/")[1].lstrip())
                                                    if((ciq_data[vip_key]["Vip Type"]).split("/")[1].lstrip() == "mip"):
                                                        pass 
                                                    elif((ciq_data[vip_key]["Vip Type"]).split("/")[1].lstrip() == vip_type):
                                                        lld_ws[cell].value = (ciq_data[vip_key]["Assigned IP(IPv4) range"]).split("/",1)[0]                          
                                                        self.cell_color(lld_ws, cell,'9400D3','ffffff')
                lld_wb.save(lld_path) 
        return(lld_path)       


    def remove_filter_data(self, *args, **kwargs):
        """
            This function is used to remove the data in every sheet from k column to last column .  
        """
        
        self.excel_file = self.excel_file_load(self.output_path)
        # lld_path = self.special_cases(args[0])
        # self.excel_file = self.excel_file_load(lld_path)

        #properties of openpyxl to format the cell
        no_fill = openpyxl.styles.PatternFill(fill_type=None)
        side = openpyxl.styles.Side(border_style=None)
        no_border = openpyxl.styles.borders.Border(left=side, right=side, top=side, bottom=side)
        
        #removing the data from k th column to last column in every row of every sheet 
        for sheet in self.excel_file:
            max_row=sheet.max_row
            max_column=sheet.max_column

            for i in range(1,max_row+1):
                # iterate over all columns
                for j in range(1,max_column+1):
                    if j >= 11:
                        cell_obj=sheet.cell(row=i,column=j)
                        if cell_obj.value != None:
                            cell_obj.value = None
                            cell_obj.fill = no_fill
                            cell_obj.border = no_border

        self.excel_file.save(self.output_path)


def main():	
        mtas = Mtas()
        mtas.create_mtas_lld(mtas)
        mtas.create_ip_sheet(mtas)
        mtas.special_cases(mtas)
        mtas.clean_redundant_column(mtas)
        # mtas.same_sheet_data(mtas)
        mtas.remove_filter_data(mtas)


if __name__=="__main__":
	main()    