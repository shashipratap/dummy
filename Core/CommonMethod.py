# -*- coding: utf-8 -*-

"""
CommonMethod.py
~~~~~~~~~~~~

This module implements the Logging package for Python. Based on PEP 282 and comments thereto in
comp.lang.python.

:copyright: (c) 2020 by Chandrayee Kumar.All Rights Reserved.
:license: Ericsson , see LICENSE for more details.

class:
CommonMethod -- responsible for calling  list of
                    resources , and managing the CIQ generation mdoules.


    methods:
    
    resource_to_execute() --

"""
import yaml
import openpyxl
import ipaddress
from math import pow
from netaddr import *
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Color, Fill, Font, PatternFill
#from Core.CustomLogger import  getLogger
from openpyxl.utils.cell import coordinate_from_string,column_index_from_string
from master_parser import parse_ciq 
from openpyxl.styles.borders import Border, Side, BORDER_THIN
__author__  = "Chandrayee Kumar <chandrayee.kumar@ericsson.com> Neeraj Sharma J <neeraj.j.sharma@ericsson.com>"
__status__  = "development"
# The following module attributes are no longer updated.
__version__ = "0.1"
__date__    = "22 April 2020"

class CommonMethod(object):
    """Raised when testcases for CommonMethod request invokes."""
    def __init__(self):
        if 'Core' in os.path.abspath("resource.yaml"): 
            self.file_path=os.path.abspath("resource.yaml").replace("\Core","")
        else:
            self.file_path=os.path.abspath("resource.yaml")  

    def __del__(self):
        classname = self.__class__.__name__
        log.info(classname +" deleted")  
    
    def resource_to_execute(self,*args,**kwargs):
        print(args[0])
        if 'Core' in os.path.abspath("resource.yaml"): 
            file_path=os.path.abspath("resource.yaml").replace("\Core","")
            
        else:
            file_path=os.path.abspath("resource.yaml")
        print(file_path)    
        resource_dict =  yaml.safe_load(open(r"C:\Users\ekcuhma\OneDrive - Ericsson AB\Local\design automation\ims-lld-automation\resource.yaml"))
        # print(args[0])
        for key,val in enumerate(resource_dict['Resources']):
            if args[0] == val: 
                return resource_dict['Resources'][val]
        
    def clean(input_string):
        temp=str(input_string).replace("('",'').replace("',)",'').replace(' ','')
        return temp    
    
    def cell_color(self, worksheet, cell,bgcolor,fontcolor):
        """Adds color to updated cells"""
        # bgcolor="openpyxl.styles.colors."+bgcolor
        # fontcolor="openpyxl.styles.colors."
        color_fill = PatternFill(end_color=bgcolor,fill_type='solid',start_color=bgcolor)
        worksheet[cell].fill = color_fill
        worksheet[cell].font = Font(color=fontcolor)       
    def border(self):
        border = Border(left=Side(border_style=BORDER_THIN,color='FF000000'),right=Side(border_style=BORDER_THIN,color='FF000000'),top=Side(border_style=BORDER_THIN, color='FF000000'),bottom=Side(border_style=BORDER_THIN,color='FF000000'))   
        return border            
    def redundant_param_name(self,*args):
        yaml_obj = self.resource_dict['Special']['Data_Redundancy']['Destination']
        p_list =[]
        for i in range(0,len(yaml_obj)):
            p_list.append([x for x in self.fvl_lld_wb[yaml_obj[i]['sheetname']][yaml_obj[i]['column']] if yaml_obj[i]['param'] == x.value])
            # print([x for x in self.fvl_lld_wb[yaml_obj[i]['sheetname']][yaml_obj[i]['column']] if yaml_obj[i]['param'] == x.value])
        return p_list
    def redundant_param_name_source(self,*args):
        yaml_obj = self.resource_dict['Special']['Data_Redundancy']['Source']
        source_list =[]
        for i in range(0,len(yaml_obj)):
            source_list.append([x for x in self.fvl_ciq_wb[yaml_obj[i]['sheetname']][yaml_obj[i]['column']] if yaml_obj[i]['param'] == x.value])
            # print([x for x in self.fvl_ciq_wb[yaml_obj[i]['sheetname']][yaml_obj[i]['column']] if yaml_obj[i]['param'] == x.value])
        return source_list    
    def ip_sheet_lookup(self,*args,**kwargs):
        # print(args[0])
        switcher = {
            # Final LLD:Customer CiQ sheet
            # 'Information':'Information',
            'Vlan Name':'Vlan Name',
            'VPN':'VPN',
            'Protocol/Mask':'Subnet size',
            'VLANID':'VLAN ID assigned by',
            'Node':'Node',
            'IPv4 Address':'Assigned IP(IPv4) range by',
            # 'Comment':'Comment',
            'IPv6 Address':'Assigned IP(Ipv6) range by',
            'Logical Connectivity Description':'Logical Connectivity Description'
                    }    
        resp = switcher.get(args[0], lambda: "Invalid Column Name")
        # print("REEEEE",resp)
        return resp
    def yaml_file_load(self,*args,**kwargs):
          
        """
               Sends a fetch request, but retrieves the
               resource_dict.Returns :dictionary :`resource_dict` object.
        
               :param args[0]: yaml file location (i.e.: self.mapping_path)
               :param **kwargs: Optional arguments that ``request`` takes.
        """

        resource_dict = yaml.safe_load(open(args[0],'r',encoding='utf-8-sig'))
        return resource_dict
    def excel_file_load(self,*args,**kwargs):
        """
               Sends a fetch request, but retrieves the
               workbook object.Returns :Object :`wb` object.
        
               :param args[0]: excel file location (i.e.: self.ciq_path,self.output_path)
               :param **kwargs: Optional arguments that ``request`` takes.
        """
        wb = load_workbook(args[0])
        # print('Inside',wb.sheetnames)
        return wb
    def validate_ciq_sheet_by_map(self,*args,**kwargs): 
        """
               Sends a search request, but retrieves the
               list of sheet object.Returns :List :`matched_sheet_name` object.
        
               :param args[0]: workbook object of CIQ (i.e.:  self.fvl_ciq_wb)
                               Dictionary object of yaml file (i.e.:  self.resource_dict)
               :param **kwargs: Optional arguments that ``request`` takes.
        """
        
        # sheet_list=self.fvl_ciq_wb.sheetnames
        sheet_list=[x.strip(' ') for x in self.fvl_ciq_wb.sheetnames]
        # print("fvl_ciq_wb",self.resource_dict['Generic'])
        
        for ciq_sheet_name in self.resource_dict['Generic']:
            # print("self.resource_dict['Generic']",ciq_sheet_name.strip())
            if ciq_sheet_name.strip() in sheet_list:
                # print("Matched",type(ciq_sheet_name))
                self.matched_sheet_name.append(ciq_sheet_name)
            else:
                #log.error("Please check your mapping file SHEET name with CIQ file SHEET name")
                #log.error("Error Code:")
                # print("Non Matched",ciq_sheet_name) 
                pass      
        return self.matched_sheet_name                                        
# def main():

#     CommonMethodObj = CommonMethod()
#     print(CommonMethodObj.resource_to_execute('client-ciq'))
    

# if __name__ == '__main__':
    
#     main()  
